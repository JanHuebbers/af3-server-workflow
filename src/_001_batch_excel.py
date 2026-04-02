#!/usr/bin/env python3
# run by: python src/_001_batch_excel.py

from pathlib import Path
from openpyxl import Workbook, load_workbook

# ---------------- user settings ----------------
EXCEL = Path("af3_input_temp.xlsx")

# Chain 1 = MLOs (9) as FASTA basenames (without .fasta) — phylogeny order
CHAIN1 = ["AtMLO1.1", "AtMLO2.1"]
print(type(CHAIN1))
# Chain 2 = EXO70s (10) as FASTA basenames (without .fasta) — phylogeny order
CHAIN2 = ["AtCAM2.1"]
print(type(CHAIN2))
START_RUNID = 1

sheet_count = len(CHAIN1) * len(CHAIN2)  # int
print(sheet_count)

# 12 seeds exactly like your example
SEEDS = [1710, 1711, 1712, 1701, 1702, 1703, 1704, 1705, 1706, 1707, 1708, 1709]

# --- chain counts (set whatever you want here) ---
CHAINA_COUNT = 1
CHAINB_COUNT = 1
CHAINC_COUNT = 0
CHAIND_COUNT = 0

HEADER = [
    "JobName","SimID","","Seed","Ions","Ions_count",
    "ChainA_count","ChainA_seq","ChainA_name",
    "ChainB_count","ChainB_seq","ChainB_name",
    "ChainC_count","ChainC_seq","ChainC_name",
    "ChainD_count","ChainD_seq","ChainD_name",
]

# ---------------- helpers ----------------
def base_name(seq_name: str) -> str:
    """Turn 'AtMLO1.1' -> 'AtMLO1' for sheet/run naming (HvMlo1.1 -> HvMlo1)."""
    return seq_name.split(".", 1)[0]

def seq_path(seq_name: str) -> str:
    """Excel path like sequences\\AtMLO1.1.fasta"""
    return fr"sequences\{seq_name}.fasta"

def make_sheet_name(runid4: str, c1_base: str, c2_base: str) -> str:
    return f"{runid4}_1{c1_base}_1{c2_base}"

def make_jobname(runid4: str, idx2: str, seed: int, c1_base: str, c2_base: str) -> str:
    # Example: 0164-02-1711_0Ions_1AtMLO11AtEXO70A1
    return f"{runid4}-{idx2}-{seed}_0Ions_{CHAINA_COUNT}{c1_base}1{c2_base}"

def main():
    # Create new workbook if file doesn't exist; otherwise load.
    if EXCEL.exists():
        wb = load_workbook(EXCEL)
    else:
        wb = Workbook()
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    runid = START_RUNID

    for c1 in CHAIN1:
        c1_base = base_name(c1)
        for c2 in CHAIN2:
            c2_base = base_name(c2)

            runid4 = f"{runid:04d}"
            sheet_name = make_sheet_name(runid4, c1_base, c2_base)

            # overwrite sheet if it exists
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

            ws = wb.create_sheet(title=sheet_name)

            # header
            for col, h in enumerate(HEADER, start=1):
                ws.cell(row=1, column=col, value=h)

            # 12 rows
            for i, seed in enumerate(SEEDS, start=1):
                idx2 = f"{i:02d}"
                jobname = make_jobname(runid4, idx2, seed, c1_base, c2_base)

                row = 1 + i
                values = [
                    jobname, runid4, idx2, seed, "Ions", 0,
                    CHAINA_COUNT, seq_path(c1), c1,
                    CHAINB_COUNT, seq_path(c2), c2,
                    CHAINC_COUNT, "", "",
                    CHAIND_COUNT, "", "",
                ]
                for col, v in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=v)

            runid += 1

    wb.save(EXCEL)
    print(f"✅ Wrote {sheet_count} sheets into {EXCEL}")

if __name__ == "__main__":
    main()